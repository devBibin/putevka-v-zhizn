import json
import logging
import os
import time
import tempfile
from copy import copy
from typing import Any, Dict, List, Optional

import django
from dotenv import load_dotenv
from django.core.files.base import ContentFile
from django.db import models, transaction
from django.utils import timezone

from openai import OpenAI
from openpyxl import load_workbook

load_dotenv()

os.environ.setdefault("DJANGO_SETTINGS_MODULE", os.getenv("DJANGO_SETTINGS_MODULE", "Putevka.settings"))
django.setup()

logger = logging.getLogger(__name__)
logging.basicConfig(level=os.getenv("LOG_LEVEL", "INFO"))

from review_by_tutor.models import Interview

try:
    from review_by_tutor.models import InterviewTemplate
except Exception:
    InterviewTemplate = None

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY") or os.getenv("GPT_TOKEN")
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
POLLING_INTERVAL = int(os.getenv("INTERVIEW_FILL_POLLING_INTERVAL", "60"))
BATCH_LIMIT = int(os.getenv("INTERVIEW_FILL_BATCH_LIMIT", "2"))


client = OpenAI(api_key=OPENAI_API_KEY)


SYSTEM_PROMPT = (
    "Ты — ассистент интервьюера. "
    "Тебе дан список полей Excel-формы интервью и транскрипт разговора. "
    "Нужно заполнить поля максимально точно по смыслу, без выдумок. "
    "Если данных нет — верни пустую строку."
)

USER_PROMPT = (
    "Верни ТОЛЬКО JSON (без текста вокруг):\n"
    "- ключи: ровно как названия полей (label) из списка\n"
    "- значения: что вписать в форму\n"
    "Правила:\n"
    "1) Не придумывай факты.\n"
    "2) Пиши кратко и по делу, но так, чтобы это было полезно интервьюеру.\n"
    "3) Если поле предполагает 'да/нет' — верни 'да' или 'нет' (если ясно из транскрипта), иначе пусто.\n"
)

from openpyxl.utils.cell import range_boundaries

def _resolve_write_cell(ws, coord: str):
    for merged in ws.merged_cells.ranges:
        if coord in merged:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged))
            return ws.cell(row=min_row, column=min_col)
    return ws[coord]

def _is_empty_like(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        s = v.strip()
        return s == "" or set(s) <= {"_", "—", "-", " "}
    return False


def extract_fields_from_workbook(template_path: str):
    wb = load_workbook(template_path)
    fields: List[Dict[str, str]] = []

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue

                label = cell.value.strip()
                if len(label) < 2:
                    continue

                if len(label) > 120:
                    continue

                right = ws.cell(row=cell.row, column=cell.column + 1)
                if _is_empty_like(right.value):
                    fields.append({
                        "sheet": ws.title,
                        "label": label,
                        "target": right.coordinate,
                    })

                down = ws.cell(row=cell.row + 1, column=cell.column)
                if _is_empty_like(down.value):
                    fields.append({
                        "sheet": ws.title,
                        "label": label,
                        "target": down.coordinate,
                    })

    return wb, fields

def ask_openai_fill(labels: List[str], transcript: str) -> Dict[str, str]:
    if not OPENAI_API_KEY:
        raise RuntimeError("OPENAI_API_KEY/GPT_TOKEN не установлен.")

    resp = client.chat.completions.create(
        model=OPENAI_MODEL,
        temperature=0,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": USER_PROMPT},
            {"role": "user", "content": f"Поля формы (labels):\n{json.dumps(labels, ensure_ascii=False)}"},
            {"role": "user", "content": f"Транскрипт:\n{transcript}"},
        ],
        response_format={"type": "json_object"},
    )

    raw = resp.choices[0].message.content or "{}"
    data = json.loads(raw)

    out: Dict[str, str] = {}
    for k in labels:
        v = data.get(k, "")
        if v is None:
            out[k] = ""
        elif isinstance(v, (dict, list)):
            out[k] = json.dumps(v, ensure_ascii=False)
        else:
            out[k] = str(v).strip()
    return out

def fill_workbook(wb, fields: List[Dict[str, str]], answers: Dict[str, str]) -> None:
    for f in fields:
        ws = wb[f["sheet"]]
        cell = _resolve_write_cell(ws, f["target"])
        label = f["label"]
        value = answers.get(label, "").strip()

        style = copy(cell._style)
        alignment = copy(cell.alignment)

        cell.value = value

        cell._style = style
        try:
            cell.alignment = alignment.copy(wrapText=True)
        except Exception:
            cell.alignment = alignment


def get_active_template_path() -> str:
    if InterviewTemplate is not None:
        obj = InterviewTemplate.objects.filter(is_active=True).order_by("-uploaded_at").first()
        if obj and getattr(obj, "file", None):
            file_field = getattr(obj, "file", None) or getattr(obj, "xlsx", None) or getattr(obj, "template", None)
            if file_field:
                try:
                    return file_field.path
                except Exception:
                    raise RuntimeError("Шаблон хранится не локально (нет .path). Нужен локальный файл или скачивание.")

    raise RuntimeError("Не найден активный Excel-шаблон: нет InterviewTemplate(is_active=True).")

def pick_interviews_to_fill(limit: int):
    qs = Interview.objects.filter(
        transcript_status="DONE",
    ).exclude(
        transcript__isnull=True
    ).exclude(
        transcript__exact=""
    ).filter(
        models.Q(filled_form__isnull=True) | models.Q(filled_form__exact="")
    ).order_by("updated_at")[:limit]
    return qs


def process_one(interview_id: int, template_path: str) -> None:
    with transaction.atomic():
        obj = Interview.objects.select_for_update().get(pk=interview_id)

        if obj.filled_form:
            return

        transcript = obj.transcript.strip()

    wb, fields = extract_fields_from_workbook(template_path)
    if not fields:
        raise RuntimeError("Не удалось распознать поля в Excel-шаблоне (эвристика не нашла пустых ячеек рядом с лейблами).")

    labels = [f["label"] for f in fields]

    answers = ask_openai_fill(labels, transcript)

    fill_workbook(wb, fields, answers)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    try:
        wb.save(tmp_path)
        with open(tmp_path, "rb") as f:
            content = f.read()

        filename = f"interview_{interview_id}_filled_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        with transaction.atomic():
            obj = Interview.objects.select_for_update().get(pk=interview_id)
            if obj.filled_form:
                return

            obj.filled_form.save(filename, ContentFile(content), save=False)
            obj.filled_uploaded_at = timezone.now()
            obj.filled_uploaded_by_id = None
            obj.save(update_fields=["filled_form", "filled_uploaded_at", "filled_uploaded_by", "updated_at"])

    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass


def fill_pending_interviews():
    template_path = get_active_template_path()
    if not template_path.lower().endswith(".xlsx"):
        raise RuntimeError("Шаблон должен быть .xlsx (openpyxl не поддерживает .xls).")

    interviews = pick_interviews_to_fill(BATCH_LIMIT)
    if not interviews.exists():
        logger.info("Нет интервью для заполнения формы.")
        return

    logger.info(f"Найдено {interviews.count()} интервью для заполнения формы.")

    for it in interviews:
        logger.info(f"Заполнение формы для Interview ID {it.id} ...")
        try:
            process_one(it.id, template_path)
            logger.info(f"  -> OK Interview ID {it.id}")
        except Exception as e:
            logger.error(f"  -> FAIL Interview ID {it.id}: {e}")


def main():
    logger.info("Запуск фонового скрипта Interview Form Filler...")
    while True:
        try:
            fill_pending_interviews()
        except Exception as e:
            logger.error(f"Критическая ошибка в главном цикле: {e}")
            time.sleep(30)

        logger.info(f"Следующая проверка через {POLLING_INTERVAL} секунд...")
        time.sleep(POLLING_INTERVAL)


if __name__ == "__main__":
    main()
