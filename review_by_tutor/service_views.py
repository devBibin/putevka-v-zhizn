from io import BytesIO

from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from review_by_tutor.services.staff_users import build_staff_users_queryset
from review_by_tutor.views import _staff_check


def _dt(value):
    if not value:
        return ""
    try:
        return timezone.localtime(value).strftime("%d.%m.%Y %H:%M")
    except Exception:
        return str(value)


def _d(value):
    if not value:
        return ""
    try:
        return value.strftime("%d.%m.%Y")
    except Exception:
        return str(value)


def _bool(value):
    if value is None:
        return ""
    return "Да" if value else "Нет"


@login_required
@user_passes_test(_staff_check)
def export_users_xlsx(request):
    qs = build_staff_users_queryset(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"

    MAX_TESTS = 1

    test_headers = []
    for i in range(1, MAX_TESTS + 1):
        test_headers += [
            f"test{i}_title",
            f"test{i}_status",
            f"test{i}_score_a",
            f"test{i}_score_b",
            f"test{i}_score_c",
            f"test{i}_percentile_a",
            f"test{i}_percentile_b",
            f"test{i}_percentile_c",
            f"test{i}_due_at",
            f"test{i}_completed_at",
        ]

    headers = [
                  # User
                  "Email",

                  "UI Фамилия",
                  "UI Имя",
                  "UI Отчество",
                  "UI Пол",
                  "UI Дата рождения",
                  "UI номер телефона",
                  "UI регион",
                  "UI город",
                  "UI адрес",

                  "Это персонал?",
                  "Дата регистрации",
                  "Дата последнего входа в систему",

                  # TelegramAccount
                  "TG username",
                  "TG имя",
                  "TG фамилия",
                  "TG language_code",
                  "TG подтверждён?",

                  # MotivationLetter
                  "ML обработано нейронкой?",
                  "ML дедлайн",
                  "ML баллы от персонала",
                  "ML рейтинг от персонала",
                  "ML комментарий на дописывание",
                  "ML дописывание запрошено в",
                  "ML дописывание запрошено от (id персонала)",
                  "ML статус",
                  "ML отправлено",
                  "ML word_count",

                  # UserInfo
                  "UI шаг отбора",
                  "UI статус анкеты",
                  "UI комментарий доработки",
                  "UI доработка запрошена в",
                  "UI аватар (ссылка)",
                  "UI status",



                  "UI название школы",
                  "UI адрес школы",
                  "UI классный руководитель",
                  "UI класс в следующем году обучения",
                  "UI профиль класса",
                  "UI планируемые экзамены",
                  "UI оценки по предметам",

                  "UI олимпиадные планы",
                  "UI планы на поступление",
                  "UI приоритетные вузы",
                  "UI направления",

                  "UI мама",
                  "UI папа",
                  "UI опекун",
                  "UI количество сиблингов",
                  "UI сиблинги",
                  "UI размер семьи",
                  "UI доход на одного члена",
                  "UI статус малоимущих",
                  "UI субсидии от государства",
                  "UI другие факторы",
                  "UI есть компьютер с интернетом",

                  "UI ссылка на вк",
                  "UI достижения",
                  "UI план подготовки",
                  "UI помощь о фонда",
                  "UI откуда услышал про программу",
                  "UI готов участвовать",

                  "UI согласен с условиями программы",
                  "UI согласен с политикой конфиденциальности",
                  "UI согласен на обработку персональных данных",
                  "UI согласен предоставить документы",

                  "UI итог от персонала",
                  "UI средние оценки за последние несколько периодов",
                  "UI семейный материальный статус",
                  "UI внутренний учебный профиль",

                  "UI многодетная семья",
                  "UI неполная семья",
                  "UI инвалидность кандидата",
                  "UI сирота или под опекой",
                  "UI потеря кормильца",
                  "UI инвалидность близкого",
                  "UI родитель-пенсионер",
                  "UI родитель на сво",
                  "UI сиблинг выпускника",

                  "UI тип населённого пункта",
                  "UI другие факты о жизненной ситуации",

                  # ScholarVideo
                  "VIDEO обзор",
                  "VIDEO баллы",
                  "VIDEO дедлайн",
                  "VIDEO ссылка на файл",
              ] + test_headers

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for user in qs.iterator(chunk_size=200):
        ui = getattr(user, "user_info", None)
        tg = getattr(user, "telegram_account", None)
        ml = getattr(user, "motivation_letter", None)
        video = getattr(user, "scholar_video", None)

        planned_exams = ""
        if ui:
            planned_exams = ", ".join(
                str(x) for x in ui.planned_exams.all()
            )

        tests = list(user.test_assignments.all())
        tests = list(user.test_assignments.all().order_by("assigned_at")[:MAX_TESTS])

        test_cells = []

        for t in tests:
            test_cells += [
                t.title,
                t.status,
                t.score_a,
                t.score_b,
                t.score_c,
                t.percentile_a,
                t.percentile_b,
                t.percentile_c,
                _dt(t.due_at),
                _dt(t.completed_at),
            ]

        missing = MAX_TESTS - len(tests)

        for _ in range(missing):
            test_cells += [""] * 10

        row = [
                  # User
                  user.email,

                  ui.last_name if ui else "",
                  ui.first_name if ui else "",
                  ui.middle_name if ui else "",
                  ui.gender if ui else "",
                  _d(ui.birth_date) if ui else "",
                  ui.phone if ui else "",
                  ui.region if ui else "",
                  ui.city if ui else "",
                  ui.address if ui else "",

                  _bool(user.is_staff),
                  _dt(user.date_joined),
                  _dt(user.last_login),

                  # TelegramAccount
                  tg.username if tg else "",
                  tg.first_name if tg else "",
                  tg.last_name if tg else "",
                  tg.language_code if tg else "",
                  _bool(tg.telegram_verified) if tg else "",

                  # MotivationLetter
                  _bool(ml.is_done) if ml else "",
                  _dt(ml.deadline_at) if ml else "",
                  ml.admin_score if ml else "",
                  ml.admin_rating if ml else "",
                  ml.revision_comment if ml else "",
                  _dt(ml.revision_requested_at) if ml else "",
                  ml.revision_requested_by_id if ml else "",
                  ml.status if ml else "",
                  _dt(ml.submitted_at) if ml else "",
                  ml.word_count() if ml else "",

                  # UserInfo
                  ui.selection_step if ui else "",
                  ui.form_status if ui else "",
                  ui.revision_comment if ui else "",
                  _dt(ui.revision_requested_at) if ui else "",
                  ui.avatar.url if ui and ui.avatar else "",
                  ui.status if ui else "",



                  ui.school_name if ui else "",
                  ui.school_address if ui else "",
                  ui.class_teacher if ui else "",
                  ui.next_year_class_digit if ui else "",
                  ui.class_profile if ui else "",
                  planned_exams,
                  ui.subject_grades if ui else "",

                  ui.olympiad_plans if ui else "",
                  ui.admission_path if ui else "",
                  ui.target_universities if ui else "",
                  ui.specializations if ui else "",

                  ui.mother if ui else "",
                  ui.father if ui else "",
                  ui.legal_guardian if ui else "",
                  ui.siblings_count if ui else "",
                  ui.siblings_info if ui else "",
                  ui.family_size if ui else "",
                  ui.income_per_member if ui else "",
                  ui.is_low_income if ui else "",
                  ui.receives_subsidy if ui else "",
                  ui.other_factors if ui else "",
                  ui.has_pc_with_internet if ui else "",

                  ui.vk if ui else "",
                  ui.achievements if ui else "",
                  ui.preparation_plan if ui else "",
                  ui.foundation_help if ui else "",
                  ui.heard_about_program if ui else "",
                  ui.willing_to_participate if ui else "",

                  _bool(ui.agree_program_conditions) if ui else "",
                  _bool(ui.agree_privacy_policy) if ui else "",
                  _bool(ui.agree_processing) if ui else "",
                  _bool(ui.agree_documents) if ui else "",

                  ui.tutor_summary if ui else "",
                  ui.avg_grade_last_period if ui else "",
                  ui.family_material_status if ui else "",
                  ui.internal_study_profile if ui else "",

                  _bool(ui.is_large_family) if ui else "",
                  _bool(ui.is_single_parent_family) if ui else "",
                  _bool(ui.has_candidate_disability) if ui else "",
                  _bool(ui.is_orphan_or_under_guardianship) if ui else "",
                  _bool(ui.has_breadwinner_loss) if ui else "",
                  _bool(ui.has_relative_disability) if ui else "",
                  _bool(ui.is_parent_pensioner) if ui else "",
                  _bool(ui.is_parent_in_svo) if ui else "",
                  _bool(ui.has_alumni_sibling) if ui else "",

                  ui.settlement_type if ui else "",
                  ui.life_situation_notes if ui else "",

                  # ScholarVideo
                  video.review if video else "",
                  video.score if video else "",
                  _dt(video.deadline_at) if video else "",
                  video.yandex_disk_path if video and video.yandex_disk_path else (video.file.url if video and video.file else ""),
              ] + test_cells

        ws.append(row)

    # Немного приводим лист в порядок
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 12,
        "B": 20,
        "C": 18,
        "D": 18,
        "E": 28,
    }

    if len(headers) != len(row):
        raise ValueError(f"Header/row mismatch: {len(headers)} != {len(row)}")

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for col_idx in range(6, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"users_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
