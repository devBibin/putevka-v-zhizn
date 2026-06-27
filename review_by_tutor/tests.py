from decimal import Decimal
from io import BytesIO
import shutil
import tempfile

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from django.test import TestCase, override_settings
from openpyxl import Workbook, load_workbook

from core.models import MotivationLetter, MotivationLetterRubricReview
from my_study.models import Course, CourseSelection, School, Subject
from review_by_tutor.models import Interview, InterviewResult, InterviewTemplate
from review_by_tutor.services.interview_xlsx import (
    _build_application_values,
    _label_key,
    build_prefilled_interview_xlsx,
    import_interview_result_xlsx,
)
from scholar_form.models import ScholarVideo, UserInfo

TEMP_MEDIA_ROOT = tempfile.mkdtemp()


@override_settings(MEDIA_ROOT=TEMP_MEDIA_ROOT)
class InterviewXlsxImportTests(TestCase):
    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(TEMP_MEDIA_ROOT, ignore_errors=True)

    def _xlsx_payload(self):
        workbook = Workbook()
        ws = workbook.active
        ws["B12"] = "Школа кандидата, номер"
        ws["D12"] = "57"
        ws["B14"] = "Как далеко школа"
        ws["D14"] = "3,5 км"
        ws["C255"] = "Комментарий"
        ws["F255"] = "Кандидат мотивирован"
        ws["C258"] = "Итоговая оценка"
        ws["F258"] = "87 баллов"

        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)
        return payload

    def test_import_interview_result_xlsx_updates_model_fields(self):
        user = get_user_model().objects.create_user(username="xlsx-import")
        interview = Interview.objects.create(user=user)
        result = InterviewResult.objects.create(interview=interview)

        updated_fields = import_interview_result_xlsx(self._xlsx_payload(), result)
        result.refresh_from_db()

        self.assertIn("school_number", updated_fields)
        self.assertIn("school_distance_km", updated_fields)
        self.assertIn("interviewer_summary", updated_fields)
        self.assertIn("interviewer_score", updated_fields)
        self.assertEqual(result.school_number, "57")
        self.assertEqual(result.school_distance_km, Decimal("3.50"))
        self.assertEqual(result.interviewer_summary, "Кандидат мотивирован")
        self.assertEqual(result.interviewer_score, 87)

    def test_import_interview_result_xlsx_prefers_candidate_answer_column_e(self):
        workbook = Workbook()
        ws = workbook.active
        ws["B12"] = "РЁРєРѕР»Р° РєР°РЅРґРёРґР°С‚Р°, РЅРѕРјРµСЂ"
        ws["D12"] = "РџРѕРґСЃРєР°Р·РєР° РёР· Р°РЅРєРµС‚С‹"
        ws["E12"] = "РћС‚РІРµС‚ РєР°РЅРґРёРґР°С‚Р°"
        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)
        user = get_user_model().objects.create_user(username="xlsx-import-column-e")
        interview = Interview.objects.create(user=user)
        result = InterviewResult.objects.create(interview=interview)

        updated_fields = import_interview_result_xlsx(payload, result)
        result.refresh_from_db()

        self.assertIn("school_number", updated_fields)
        self.assertEqual(result.school_number, "РћС‚РІРµС‚ РєР°РЅРґРёРґР°С‚Р°")

    def test_upload_interview_template_saves_file_and_updates_result(self):
        staff = get_user_model().objects.create_user(username="staff", is_staff=True)
        user = get_user_model().objects.create_user(username="candidate")
        interview = Interview.objects.create(user=user)
        InterviewResult.objects.create(interview=interview)
        upload = SimpleUploadedFile(
            "filled.xlsx",
            self._xlsx_payload().getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        self.client.force_login(staff)
        response = self.client.post(
            reverse("interview_detail", args=[user.id]),
            {"op": "upload_interview_template", "filled_template": upload},
        )

        self.assertEqual(response.status_code, 302)
        interview.refresh_from_db()
        result = InterviewResult.objects.get(interview=interview)
        self.assertTrue(interview.filled_template.name.endswith(".xlsx"))
        self.assertIsNotNone(interview.filled_uploaded_at)
        self.assertEqual(result.school_number, "57")
        self.assertEqual(result.interviewer_score, 87)

    def test_prefilled_interview_xlsx_does_not_use_interview_result_values(self):
        workbook = Workbook()
        ws = workbook.active
        ws["B12"] = "Школа кандидата, номер"
        ws["C255"] = "Комментарий"
        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)
        InterviewTemplate.objects.create(
            title="Template",
            file=SimpleUploadedFile(
                "template.xlsx",
                payload.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        )
        user = get_user_model().objects.create_user(username="prefill-no-result")
        interview = Interview.objects.create(user=user)
        InterviewResult.objects.create(
            interview=interview,
            school_number="57",
            interviewer_summary="Кандидат мотивирован",
        )

        result_payload = build_prefilled_interview_xlsx(user, interview)
        result_workbook = load_workbook(BytesIO(result_payload), data_only=True)
        result_ws = result_workbook.active

        self.assertIsNone(result_ws["D12"].value)
        self.assertIsNone(result_ws["F255"].value)

    def test_interview_detail_can_advance_candidate_after_interview(self):
        staff = get_user_model().objects.create_user(username="advance-staff", is_staff=True)
        user = get_user_model().objects.create_user(username="advance-candidate")
        user_info = UserInfo.objects.create(
            user=user,
            selection_step=UserInfo.SelectionStep.INTERVIEW_PREP,
        )

        self.client.force_login(staff)
        response = self.client.post(
            reverse("interview_detail", args=[user.id]),
            {"op": "advance_selection_step"},
        )

        self.assertEqual(response.status_code, 302)
        user_info.refresh_from_db()
        self.assertEqual(user_info.selection_step, UserInfo.SelectionStep.AFTER_INTERVIEW)

    def test_after_interview_candidate_sees_waiting_message(self):
        user = get_user_model().objects.create_user(username="after-interview-candidate")
        UserInfo.objects.create(
            user=user,
            selection_step=UserInfo.SelectionStep.AFTER_INTERVIEW,
        )

        self.client.force_login(user)
        response = self.client.get(reverse("preparation"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Спасибо большое, что уделил(-а) время собеседованию!")
        self.assertContains(response, "Объявить результаты отбора мы планируем ближе к концу августа.")
        self.assertContains(response, "Я сообщил(-а) родителям/ законным представителям")
    def test_after_interview_candidate_can_save_checkboxes(self):
        user = get_user_model().objects.create_user(username="after-interview-checkboxes")
        user_info = UserInfo.objects.create(
            user=user,
            selection_step=UserInfo.SelectionStep.AFTER_INTERVIEW,
        )

        self.client.force_login(user)
        response = self.client.post(
            reverse("preparation"),
            {
                "after_interview_parents_notified": "on",
                "after_interview_documents_ready": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        user_info.refresh_from_db()
        self.assertTrue(user_info.after_interview_parents_notified)
        self.assertTrue(user_info.after_interview_documents_ready)


        response = self.client.post(reverse("preparation"), {})

        self.assertEqual(response.status_code, 302)
        user_info.refresh_from_db()
        self.assertTrue(user_info.after_interview_parents_notified)
        self.assertTrue(user_info.after_interview_documents_ready)

    def test_interview_xlsx_uses_human_letter_comment_for_post_letter_questions(self):
        user = get_user_model().objects.create_user(username="letter-comment-match")
        interview = Interview.objects.create(user=user)
        letter = MotivationLetter.objects.create(
            user=user,
            letter_text="Текст письма",
            admin_rating="Комментарий от человека",
        )
        MotivationLetterRubricReview.objects.create(
            letter=letter,
            reviewer_comment="Комментарий из рубрики",
        )

        values = _build_application_values(user, interview, None)

        self.assertEqual(
            values["Вопросы на собеседование после мотписьма"],
            "Комментарий от человека",
        )

    def test_interview_xlsx_combines_letter_and_video_rubric_fields_in_extracts(self):
        user = get_user_model().objects.create_user(username="combined-extracts")
        interview = Interview.objects.create(user=user)
        letter = MotivationLetter.objects.create(user=user, letter_text="Letter text")
        MotivationLetterRubricReview.objects.create(
            letter=letter,
            specialty="Computer science",
            preferred_universities="TSU",
            motivation="Strong motivation",
            help_criticality="Needs support",
        )
        ScholarVideo.objects.create(
            user=user,
            score=85,
            review="Clear video review",
            online_school_selected_courses="Math course",
            online_school_choice_reason="Needs structure",
            schedule_realistic_assessment="Realistic schedule",
            schedule_interview_questions="How will you keep pace?",
        )

        values = _build_application_values(user, interview, None)
        extracts = values["Выписки из мотивационного письма"]

        self.assertIn("Computer science", extracts)
        self.assertIn("Strong motivation", extracts)
        self.assertIn("Clear video review", extracts)
        self.assertIn("Math course", extracts)
        self.assertIn("Realistic schedule", extracts)
        self.assertIn("How will you keep pace?", extracts)

    def test_prefilled_interview_xlsx_writes_c_label_values_to_d(self):
        workbook = Workbook()
        ws = workbook.active
        ws["C12"] = "Выписки из мотписьма и видеовизитки"
        ws["C13"] = "По каким предметам планируешь сдавать ЕГЭ?"
        ws["H14"] = "Выбранные курсы и онлайн-школы"
        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)
        InterviewTemplate.objects.create(
            title="Template C labels",
            file=SimpleUploadedFile(
                "template-c-labels.xlsx",
                payload.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        )
        user = get_user_model().objects.create_user(username="prefill-c-labels")
        interview = Interview.objects.create(user=user)
        subject = Subject.objects.create(name="Информатика", slug="informatics")
        info = UserInfo.objects.create(user=user)
        info.planned_exams.add(subject)
        letter = MotivationLetter.objects.create(user=user, letter_text="Letter text")
        MotivationLetterRubricReview.objects.create(
            letter=letter,
            specialty="ИТ",
            motivation="Высокая мотивация",
        )
        ScholarVideo.objects.create(
            user=user,
            review="Хороший отзыв",
            online_school_selected_courses="Профильная математика",
        )

        result_payload = build_prefilled_interview_xlsx(user, interview)
        result_workbook = load_workbook(BytesIO(result_payload), data_only=True)
        result_ws = result_workbook.active

        self.assertIn("ИТ", result_ws["D12"].value)
        self.assertIn("Хороший отзыв", result_ws["D12"].value)
        self.assertEqual(result_ws["D13"].value, "Информатика")
        self.assertEqual(result_ws["H15"].value, "Профильная математика")

    def test_prefilled_interview_xlsx_fuzzy_matches_c_labels_to_d(self):
        workbook = Workbook()
        ws = workbook.active
        ws["C12"] = "Есть ли дома компьютер или ноутбук с интернетом?"
        ws["C13"] = "химия"
        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)
        InterviewTemplate.objects.create(
            title="Template fuzzy C labels",
            file=SimpleUploadedFile(
                "template-fuzzy-c-labels.xlsx",
                payload.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        )
        user = get_user_model().objects.create_user(username="prefill-fuzzy-c-labels")
        interview = Interview.objects.create(user=user)
        UserInfo.objects.create(
            user=user,
            has_pc_with_internet="Есть ноутбук и стабильный интернет",
        )

        result_payload = build_prefilled_interview_xlsx(user, interview)
        result_workbook = load_workbook(BytesIO(result_payload), data_only=True)
        result_ws = result_workbook.active

        self.assertEqual(result_ws["D12"].value, "Есть ноутбук и стабильный интернет")
        self.assertIsNone(result_ws["D13"].value)

    def test_interview_xlsx_label_key_matches_cyrillic_template_labels(self):
        self.assertEqual(
            _label_key("Как ты планируешь свою подготовку к поступлению на следующий год?"),
            _label_key("Как_ты_планируешь_свою_подготовку_к_поступлению_на_следующий_год_"),
        )

    def test_interview_xlsx_uses_selected_courses_from_scholar_video(self):
        user = get_user_model().objects.create_user(username="video-selected-courses")
        UserInfo.objects.create(
            user=user,
            preparation_plan="Общий план подготовки",
        )
        video = ScholarVideo.objects.create(
            user=user,
            online_school_selected_courses="Курсы из видеовизитки",
        )
        interview = Interview.objects.create(user=user)

        values = _build_application_values(user, interview, None)

        self.assertEqual(
            values["Выбранные курсы и онлайн-школы"],
            video.online_school_selected_courses,
        )

    def test_interview_xlsx_falls_back_to_candidate_online_school_course(self):
        user = get_user_model().objects.create_user(username="video-course-fallback")
        video = ScholarVideo.objects.create(
            user=user,
            online_school_course="Candidate course text",
            online_school_selected_courses="",
        )
        interview = Interview.objects.create(user=user)

        values = _build_application_values(user, interview, None)

        self.assertEqual(
            values["Выбранные курсы и онлайн-школы"],
            video.online_school_course,
        )

    def test_interview_xlsx_does_not_add_study_course_selections_to_preparation_plan(self):
        user = get_user_model().objects.create_user(username="study-course-not-in-xlsx")
        UserInfo.objects.create(
            user=user,
            preparation_plan="General preparation plan",
        )
        subject = Subject.objects.create(name="Math", slug="math")
        school = School.objects.create(name="Online School")
        course = Course.objects.create(school=school, subject=subject, title="Exam prep")
        CourseSelection.objects.create(user=user, course=course, motivation="Need course")
        interview = Interview.objects.create(user=user)

        values = _build_application_values(user, interview, None)

        self.assertEqual(
            values["Как_ты_планируешь_свою_подготовку_к_поступлению_на_следующий_год_"],
            "General preparation plan",
        )
        self.assertNotIn("Online School", values["Как_ты_планируешь_свою_подготовку_к_поступлению_на_следующий_год_"])

    def test_interview_xlsx_uses_video_review_for_video_internal_note(self):
        user = get_user_model().objects.create_user(username="video-review-note")
        video = ScholarVideo.objects.create(
            user=user,
            review="Отзыв из видеовизитки",
            online_school_interview_questions="Вопросы из видеовизитки",
        )
        interview = Interview.objects.create(user=user)

        values = _build_application_values(user, interview, None)

        self.assertEqual(
            values["Внутренняя заметка по видеовизитке"],
            video.review,
        )

    def test_interview_xlsx_leaves_interviewer_questions_empty(self):
        user = get_user_model().objects.create_user(username="empty-interviewer-questions")
        ScholarVideo.objects.create(
            user=user,
            online_school_interview_questions="Вопрос по курсам",
            schedule_interview_questions="Вопрос по графику",
        )
        interview = Interview.objects.create(
            user=user,
            notes="Заметка интервьюера",
        )

        values = _build_application_values(user, interview, None)

        self.assertEqual(values["Вопросы на собеседование от интервьюера"], "")
