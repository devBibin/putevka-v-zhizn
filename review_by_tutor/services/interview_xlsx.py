from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
import re

from django.conf import settings
from django.db import models
from django.utils.dateparse import parse_date, parse_datetime
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from core.models import MotivationLetter
from my_study.models import UniversityPriority
from review_by_tutor.models import InterviewTemplate, TestAssignment
from scholar_form.models import ScholarVideo, UserInfo


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Да" if value else "Нет"
    if hasattr(value, "strftime"):
        return timezone.localtime(value).strftime("%d.%m.%Y %H:%M") if hasattr(value, "tzinfo") else value.strftime("%d.%m.%Y")
    return str(value).strip()


def _join(items) -> str:
    return "\n".join(_text(item) for item in items if _text(item))


def _labeled_value(label: str, value) -> str:
    value_text = _text(value)
    return f"{label}: {value_text}" if value_text else ""


def _build_interview_extracts(rubric, video) -> str:
    video_selected_courses = _field(video, "online_school_selected_courses") or _field(video, "online_school_course")
    return _join([
        _labeled_value("Мотписьмо - специальность", _field(rubric, "specialty")),
        _labeled_value("Мотписьмо - предпочитаемые вузы", _field(rubric, "preferred_universities")),
        _labeled_value("Мотписьмо - мотивация", _field(rubric, "motivation")),
        _labeled_value("Мотписьмо - критичность помощи", _field(rubric, "help_criticality")),
        _labeled_value("Видеовизитка - балл", _field(video, "score")),
        _labeled_value("Видеовизитка - отзыв", _field(video, "review")),
        _labeled_value("Видеовизитка - опыт онлайн-школ", _field(video, "online_school_prior_experience")),
        _labeled_value("Видеовизитка - наблюдения по онлайн-школам", _field(video, "online_school_observations")),
        _labeled_value("Видеовизитка - выбранные курсы", video_selected_courses),
        _labeled_value("Видеовизитка - обоснование выбора курсов", _field(video, "online_school_choice_reason")),
        _labeled_value("Видеовизитка - дополнительная помощь по курсам", _field(video, "online_school_extra_help")),
        _labeled_value("Видеовизитка - использование материалов по курсам", _field(video, "online_school_used_materials")),
        _labeled_value("Видеовизитка - вопросы по курсам", _field(video, "online_school_interview_questions")),
        _labeled_value("Видеовизитка - школьный день", _field(video, "schedule_school_day")),
        _labeled_value("Видеовизитка - домашняя работа", _field(video, "schedule_homework_time")),
        _labeled_value("Видеовизитка - кружки/спорт/активности", _field(video, "schedule_activities_time")),
        _labeled_value("Видеовизитка - подготовка к ЕГЭ/олимпиадам", _field(video, "schedule_exam_prep_time")),
        _labeled_value("Видеовизитка - отдых", _field(video, "schedule_rest_time")),
        _labeled_value("Видеовизитка - выходной день", _field(video, "schedule_weekend_day")),
        _labeled_value("Видеовизитка - реалистичность графика", _field(video, "schedule_realistic_assessment")),
        _labeled_value("Видеовизитка - дополнительная помощь по графику", _field(video, "schedule_extra_help")),
        _labeled_value("Видеовизитка - использование материалов по графику", _field(video, "schedule_used_materials")),
        _labeled_value("Видеовизитка - вопросы по графику", _field(video, "schedule_interview_questions")),
    ])


def _repair_mojibake(value: str) -> str:
    try:
        repaired = value.encode("cp1251").decode("utf-8")
    except UnicodeError:
        return value
    return repaired if any("а" <= ch <= "я" or ch == "ё" for ch in repaired.lower()) else value


def _label_key(value) -> str:
    text = _repair_mojibake(_text(value)).lower().replace("ё", "е")
    return re.sub(r"[^0-9a-zа-я]+", "", text)


def _label_words(value) -> set[str]:
    text = _repair_mojibake(_text(value)).lower().replace("ё", "е")
    words = set(re.findall(r"[0-9a-zа-я]+", text))
    stop_words = {
        "есть", "если", "какие", "какой", "какая", "какую", "каким", "каких",
        "почему", "когда", "куда", "куда", "тебя", "тебе", "ты", "твоей",
        "твои", "свои", "свою", "свое", "свой", "ли", "или", "для", "при",
        "наличии", "которые", "которым", "которых", "последние",
        "фонд", "фонда", "программа", "программе", "программы",
    }
    return {word for word in words if len(word) >= 4 and word not in stop_words}


def _field(obj, name: str) -> str:
    if not obj:
        return ""
    display = getattr(obj, f"get_{name}_display", None)
    if callable(display):
        value = display()
        if value:
            return _text(value)
    return _text(getattr(obj, name, ""))


def _safe_set(ws, coord: str, value):
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        for merged in ws.merged_cells.ranges:
            if coord in merged:
                cell = ws.cell(merged.min_row, merged.min_col)
                break
    cell.value = value or None


def _safe_get(ws, coord: str):
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        for merged in ws.merged_cells.ranges:
            if coord in merged:
                cell = ws.cell(merged.min_row, merged.min_col)
                break
    return cell.value


def _latest_completed_test(user):
    return (
        TestAssignment.objects
        .filter(user=user)
        .order_by("-completed_at", "-assigned_at", "-id")
        .first()
    )


def _template_path() -> Path:
    template = InterviewTemplate.objects.filter(is_active=True).order_by("-uploaded_at").first()
    if template and template.file:
        return Path(template.file.path)

    for filename in ("Собеседование 2026.xlsx", "Шаблон 2025.xlsx"):
        root_template = Path(settings.BASE_DIR) / filename
        if root_template.exists():
            return root_template

    raise FileNotFoundError("Interview XLSX template not found")


def _add_value_aliases(values: dict[str, str], aliases: dict[str, str]) -> dict[str, str]:
    aliases = {
        **aliases,
        "\u041a\u0430\u043a\u0443\u044e \u043f\u043e\u043c\u043e\u0449\u044c \u0442\u044b \u0445\u043e\u0447\u0435\u0448\u044c \u043f\u043e\u043b\u0443\u0447\u0438\u0442\u044c \u043e\u0442 \u0424\u043e\u043d\u0434\u0430?": "\u041a\u0430\u043a\u0443\u044e_\u043f\u043e\u043c\u043e\u0449\u044c_\u0442\u044b_\u043e\u0436\u0438\u0434\u0430\u0435\u0448\u044c_\u043e\u0442_\u0424\u043e\u043d\u0434\u0430",
        "\u041a\u0430\u043a \u0442\u044b \u0443\u0437\u043d\u0430\u043b(-\u0430) \u043e \u043f\u0440\u043e\u0433\u0440\u0430\u043c\u043c\u0435 \u0424\u043e\u043d\u0434\u0430?": "\u041a\u0430\u043a_\u0442\u044b_\u0443\u0437\u043d\u0430\u043b_\u043e_\u043f\u0440\u043e\u0433\u0440\u0430\u043c\u043c\u0435",
        "\u0423\u0447\u0435\u0431\u043d\u044b\u0439 \u043f\u0440\u043e\u0444\u0438\u043b\u044c (\u0432\u043d\u0443\u0442\u0440\u0435\u043d\u043d\u0438\u0439)": "\u041a\u0430\u0442\u0435\u0433\u043e\u0440\u0438\u044f",
        "\u041c\u043d\u043e\u0433\u043e\u0434\u0435\u0442\u043d\u043e\u0441\u0442\u044c (\u0432\u043d\u0443\u0442\u0440\u0435\u043d\u043d\u0435\u0435)": "\u041c\u043d\u043e\u0433\u043e\u0434\u0435\u0442\u043d\u0430\u044f \u0441\u0435\u043c\u044c\u044f",
        "\u0412\u043e\u043f\u0440\u043e\u0441 5 (\u0442\u0440\u0430\u0435\u043a\u0442\u043e\u0440\u0438\u044f \u043f\u043e\u0441\u0442\u0443\u043f\u043b\u0435\u043d\u0438\u044f)": "\u0412\u043e\u043f\u0440\u043e\u0441 5 (\u043f\u043e\u0434\u0433\u043e\u0442\u043e\u0432\u043a\u0430 \u0432 \u0441\u043b\u0435\u0434\u0443\u044e\u0449\u0435\u043c \u0433\u043e\u0434\u0443)",
        "\u0412\u043e\u043f\u0440\u043e\u0441 6 (\u043f\u043e\u0434\u0433\u043e\u0442\u043e\u0432\u043a\u0430 \u0432 \u0441\u043b\u0435\u0434\u0443\u044e\u0449\u0435\u043c \u0433\u043e\u0434\u0443)": "\u0412\u043e\u043f\u0440\u043e\u0441 5 (\u043f\u043e\u0434\u0433\u043e\u0442\u043e\u0432\u043a\u0430 \u0432 \u0441\u043b\u0435\u0434\u0443\u044e\u0449\u0435\u043c \u0433\u043e\u0434\u0443)",
        "\u0412\u043e\u043f\u0440\u043e\u0441 7 (\u0437\u043d\u0430\u0447\u0438\u043c\u043e\u0441\u0442\u044c \u0412\u041e)": "\u0412\u043e\u043f\u0440\u043e\u0441 6 (\u0437\u043d\u0430\u0447\u0438\u043c\u043e\u0441\u0442\u044c \u0412\u041e)",
        "\u0412\u043e\u043f\u0440\u043e\u0441 8 (\u043a\u0440\u0438\u0442\u0438\u0447\u043d\u043e\u0441\u0442\u044c \u043f\u043e\u043c\u043e\u0449\u0438)": "\u0412\u043e\u043f\u0440\u043e\u0441 7 (\u043a\u0440\u0438\u0442\u0438\u0447\u043d\u043e\u0441\u0442\u044c \u043f\u043e\u043c\u043e\u0449\u0438)",
    }
    result = dict(values)
    for label, value in values.items():
        result.setdefault(_repair_mojibake(label), value)

    for label, source_label in aliases.items():
        value = result.get(source_label, result.get(_repair_mojibake(source_label), ""))
        result[label] = value
        result.setdefault(_repair_mojibake(label), value)
    return result


def _build_application_values(user, interview, result=None) -> dict[str, str]:
    info = UserInfo.objects.filter(user=user).first()
    letter = MotivationLetter.objects.filter(user=user).first()
    rubric = getattr(letter, "rubric_review", None) if letter else None
    video = ScholarVideo.objects.filter(user=user).first()
    test = _latest_completed_test(user)

    planned_exams = ""
    if info:
        planned_exams = _join(info.planned_exams.values_list("name", flat=True))

    universities = _join(
        f"{u.priority}. {u.university}{', ' + u.city if u.city else ''}{' - ' + u.specialty if u.specialty else ''}"
        for u in UniversityPriority.objects.filter(user=user).order_by("priority")
    )
    full_name = _join([_field(info, "last_name"), _field(info, "first_name"), _field(info, "middle_name")])
    if not full_name:
        full_name = user.get_full_name() or user.username

    one_parent_not_working = "Да" if info and info.is_single_parent_family else ""
    both_parents_not_working = ""

    values = {
        "ФИО": full_name.replace("\n", " "),
        "Дата_рождения": _field(info, "birth_date"),
        "Телефон": _field(info, "phone") or getattr(user, "email", ""),
        "Email": _field(info, "email") or user.email,
        "Город_область_регион": _join([_field(info, "region"), _field(info, "city")]),
        "Наименование_и_номер_школы": _field(info, "school_name"),
        "В какой класс ты пойдешь в следующем году?": _field(info, "next_year_class_digit"),
        "Профиль_класса_при_наличии": _field(info, "class_profile"),
        "Предметы_по_которым_планируешь_сдавать_ЕГЭ": planned_exams,
        "Средний_балл_за_последний_учебный_год_по_предметам": _field(info, "subject_grades") or _field(info, "avg_grade_last_period"),
        "Планируешь_ли_ты_участвовать_в_олимпиадах": _field(info, "olympiad_plans"),
        "Ты_планируешь_поступать_по_ЕГЭ_или_олимпиадам_": _field(info, "admission_path"),
        "Список_наиболее_приоритетных_ВУЗов_в_которые_планируешь_поступать_не_более_5": universities or _field(info, "target_universities"),
        "Список_специальностей_которые_рассматриваешь_для_поступления": _field(info, "specializations"),
        "Мать": _field(info, "mother"),
        "Отец": _field(info, "father"),
        "Иной_законный_представитель": _field(info, "legal_guardian"),
        "Количество_братьев_и_сестёр": _field(info, "siblings_count"),
        "Братья_и_сёстры": _field(info, "siblings_info"),
        "Количественный_состав_семьи_": _field(info, "family_size"),
        "Среднемесячный_доход_на_1_члена_семьи_за_последние_12_месяцев_руб_": _field(info, "income_per_member"),
        "Имеет_ли_семья_статус_малоимущей": _field(info, "is_low_income"),
        "Получает_ли_семья_субсидии_от_государства_": _field(info, "receives_subsidy"),
        "Есть_какие-либо_иные_обстоятельства_о_которых_ты_хотел-a_бы_сообщить": _join([_field(info, "other_factors"), _field(info, "life_situation_notes")]),
        "Есть_ли_у_тебя_дома_компьютер_с_доступом_в_интернет": _field(info, "has_pc_with_internet"),
        "Кратко_опиши_свои_достижения_за_последние_два_года": _field(info, "achievements"),
        "Как_ты_планируешь_свою_подготовку_к_поступлению_на_следующий_год_": _field(info, "preparation_plan"),
        "Какую_помощь_ты_ожидаешь_от_Фонда": _field(info, "foundation_help"),
        "Как_ты_узнал_о_программе": _field(info, "heard_about_program"),
        "Какую помощь ты хочешь получить от Фонда?": _field(info, "foundation_help"),
        "Как ты узнал(-а) о программе Фонда?": _field(info, "heard_about_program"),
        "Готов(-а) ли ты активно принимать участие в программе Фонда?": _field(info, "willing_to_participate"),
        "Учебный профиль (внутренний)": _field(info, "internal_study_profile"),
        "Гуманитарий": "Да" if _field(info, "internal_study_profile") == "Гуманитарий" else "",
        "Инвалидность ребенка": _field(info, "has_candidate_disability"),
        "Многодетная семья": _field(info, "is_large_family"),
        "Неполная семья": _field(info, "is_single_parent_family"),
        "Сирота / под опекой": _field(info, "is_orphan_or_under_guardianship"),
        "Многодетность (внутреннее)": _field(info, "is_large_family"),
        "Неполная семья (внутреннее)": _field(info, "is_single_parent_family"),
        "Инвалидность кандидата (внутреннее)": _field(info, "has_candidate_disability"),
        "Сирота / под опекой (внутреннее)": _field(info, "is_orphan_or_under_guardianship"),
        "Потеря кормильца (внутреннее)": _field(info, "has_breadwinner_loss"),
        "Инвалидность близкого (внутреннее)": _field(info, "has_relative_disability"),
        "Родитель пенсионер (внутреннее)": _field(info, "is_parent_pensioner"),
        "Родитель на СВО (внутреннее)": _field(info, "is_parent_in_svo"),
        "Сиблинг выпускника (внутреннее)": _field(info, "has_alumni_sibling"),
        "Особенность жизненной ситуации (внутреннее)": _field(info, "life_situation_notes"),
        "Один из родителей не работает": one_parent_not_working,
        "Оба родителя не работают": both_parents_not_working,
        "Родитель пенсионер / инвалид": _field(info, "is_parent_pensioner"),
        "Другое важное": _join([_field(info, "other_factors"), _field(info, "family_material_status")]),
        "Числовой (процентиль)": _field(test, "numeric_percentile"),
        "Числовой (грейд)": _field(test, "numeric_grade"),
        "Вербальный (процентиль)": _field(test, "verbal_percentile"),
        "Вербальный (грейд)": _field(test, "verbal_grade"),
        "Логический (процентиль)": _field(test, "logical_percentile"),
        "Логический (грейд)": _field(test, "logical_grade"),
        "Дата получения мотивационного письма": _field(letter, "submitted_at"),
        "Число слов": _field(rubric, "word_count") or (letter.word_count() if letter else ""),
        "Вопрос 2 (специальность)": _field(rubric, "specialty_choice_score"),
        "Вопрос 3 (ВУЗ)": _field(rubric, "university_choice_score"),
        "Вопрос 4 (текущая подготовка)": _field(rubric, "current_preparation_score"),
        "Вопрос 5 (траектория поступления)": _field(rubric, "admission_trajectory_score"),
        "Вопрос 5 (подготовка в следующем году)": _field(rubric, "next_year_preparation_score"),
        "Вопрос 6 (подготовка в следующем году)": _field(rubric, "next_year_preparation_score"),
        "Вопрос 6 (значимость ВО)": _field(rubric, "higher_education_value_score"),
        "Вопрос 7 (значимость ВО)": _field(rubric, "higher_education_value_score"),
        "Вопрос 7 (критичность помощи)": _field(rubric, "support_criticality_score"),
        "Вопрос 8 (критичность помощи)": _field(rubric, "support_criticality_score"),
        "Последовательность": _field(rubric, "composition_penalty"),
        "Точность и выразительность": _field(rubric, "style_penalty"),
        "Орфография": _field(rubric, "orthography_penalty"),
        "Пунктуация": _field(rubric, "syntax_penalty"),
        "Итоговый балл за мотивационное письмо": _field(rubric, "total_score") or _field(letter, "admin_score"),
        "Предварительная оценка мотивации": _field(rubric, "motivation"),
        "Предварительная оценка критичности": _field(rubric, "help_criticality"),
        "Олимпиадник": _field(rubric, "olympiads"),
        "Про олимпиады": _field(info, "olympiad_plans"),
        "Категория": _field(info, "internal_study_profile"),
        "Оценка амбиций": _field(rubric, "justification"),
        "Результаты проверки (примеры ошибок)": _field(rubric, "justification"),
        "Выписки из мотивационного письма": _build_interview_extracts(rubric, video),
        "Вопросы на собеседование после мотписьма": _field(letter, "admin_rating"),
        "Вопросы на собеседование от интервьюера": "",
        "Выбранные курсы из видеовизитки": _field(video, "online_school_selected_courses") or _field(video, "online_school_course"),
        "Отзыв по видеовизитке": _field(video, "review"),
    }
    return _add_value_aliases(values, {
        "Дата рождения": "Дата_рождения",
        "Субъект, населенный пункт, адрес": "Город_область_регион",
        "Наименование и номер школы": "Наименование_и_номер_школы",
        "В какой класс пойдешь в 2026/2027 уч.году?": "В какой класс ты пойдешь в следующем году?",
        "Профиль класса при наличии": "Профиль_класса_при_наличии",
        "Предметы, по которым планируешь сдавать ЕГЭ": "Предметы_по_которым_планируешь_сдавать_ЕГЭ",
        "Средний балл по профильным предметам за последние 2 отчетных периода": "Средний_балл_за_последний_учебный_год_по_предметам",
        "Средний балл успеваемости за последний отчетный период": "Средний_балл_за_последний_учебный_год_по_предметам",
        "Планы участия в олимпиадах": "Планируешь_ли_ты_участвовать_в_олимпиадах",
        "Ты планируешь поступать по ЕГЭ или олимпиадам?": "Ты_планируешь_поступать_по_ЕГЭ_или_олимпиадам_",
        "Приоритетные вузы": "Список_наиболее_приоритетных_ВУЗов_в_которые_планируешь_поступать_не_более_5",
        "Интересующие специальности для поступления": "Список_специальностей_которые_рассматриваешь_для_поступления",
        "Иной законный представитель": "Иной_законный_представитель",
        "Количество братьев и сестер": "Количество_братьев_и_сестёр",
        "Братья и сестры": "Братья_и_сёстры",
        "Количественный состав семьи": "Количественный_состав_семьи_",
        "Среднемесячный доход на 1 члена семьи за последние 12 месяцев (руб.)": "Среднемесячный_доход_на_1_члена_семьи_за_последние_12_месяцев_руб_",
        "Имеет ли семья статус малоимущей?": "Имеет_ли_семья_статус_малоимущей",
        "Получает ли семья субсидии от государства?": "Получает_ли_семья_субсидии_от_государства_",
        "Какие-либо иные обстоятельства, о которых ты хотел(-а) бы сообщить": "Есть_какие-либо_иные_обстоятельства_о_которых_ты_хотел-a_бы_сообщить",
        "Есть ли у тебя компьютер/ноутбук с доступом в интернет?": "Есть_ли_у_тебя_дома_компьютер_с_доступом_в_интернет",
        "Кратко опиши свои достижения за последние два года": "Кратко_опиши_свои_достижения_за_последние_два_года",
        "Как бы ты оценил(-а) материальное положение вашей семьи?": "Другое важное",
        "Как бы ты описал(-а) свою подготовку после школы?": "Как_ты_планируешь_свою_подготовку_к_поступлению_на_следующий_год_",
        "Какую помощь ты ожидаешь от Фонда?": "Какую_помощь_ты_ожидаешь_от_Фонда",
        "Как ты узнал(-а) о программе?": "Как_ты_узнал_о_программе",
        "Инвалидность кандидата (внутреннее)": "Инвалидность ребенка",
        "Многодетная семья (внутреннее)": "Многодетная семья",
        "Неполная семья (внутреннее)": "Неполная семья",
        "Сирота / под опекой (внутреннее)": "Сирота / под опекой",
        "Родитель пенсионер (внутреннее)": "Родитель пенсионер / инвалид",
        "Дата получения мотивационного письма": "Дата получения мотивационного письма",
        "Результаты проверки (примеры ошибок)": "Оценка амбиций",
        "Внутренняя заметка по мотивационному письму": "Вопросы на собеседование после мотписьма",
        "Выбранные курсы и онлайн-школы": "Выбранные курсы из видеовизитки",
        "Внутренняя заметка по видеовизитке": "Отзыв по видеовизитке",
    })


INTERVIEW_RESULT_ROWS = {
    12: "school_number",
    13: "school_type",
    14: "school_distance_km",
    15: "school_specialization",
    16: "school_students_total",
    17: "school_left_after_9_est",
    18: "school_students_11",
    19: "class_profile",
    20: "has_ege_teachers_all",
    33: "triples_reason",
    34: "favorite_teacher",
    35: "favorite_subject",
    36: "has_computer_lab",
    37: "olympiads_frequency",
    38: "clubs_info",
    46: "olympiad_support_by_school",
    47: "other_school_notes",
    49: "aims_medal",
    50: "admission_way",
    51: "ege_subjects",
    63: "olympiad_experience",
    75: "other_support_needed",
    99: "had_tutor",
}

BELOW_VALUE_LABELS = {
    "Результаты проверки (примеры ошибок)",
    "Внутренняя заметка по мотивационному письму",
    "Выбранные курсы и онлайн-школы",
    "Внутренняя заметка по видеовизитке",
    "Вопросы на собеседование от интервьюера",
}

INTERVIEW_RESULT_MARKERS = {
    "school_number": ("школа кандидата", "номер"),
    "school_type": ("тип школы",),
    "school_distance_km": ("как далеко школа",),
    "school_specialization": ("специализация",),
    "school_students_total": ("сколько учеников в школе",),
    "school_left_after_9_est": ("покинуло школу после 9",),
    "school_students_11": ("учеников в 11",),
    "class_profile": ("профиль класса",),
    "has_ege_teachers_all": ("профильные учителя", "егэ"),
    "triples_reason": ("если есть тройки",),
    "favorite_teacher": ("преподаватель самый любимый",),
    "favorite_subject": ("предмет самый любимый",),
    "has_computer_lab": ("компьютер", "класс"),
    "olympiads_frequency": ("часто", "проводятся", "олимпиады"),
    "clubs_info": ("кружков", "внеклассные"),
    "olympiad_support_by_school": ("олимпиадному движению",),
    "other_school_notes": ("другие сведения о школе",),
    "aims_medal": ("медаль", "аттестат"),
    "admission_way": ("поступать по егэ или олимпиадам",),
    "ege_subjects": ("по каким предметам", "планируешь сдавать", "егэ"),
    "olympiad_experience": ("участвовал ли", "олимпиадах"),
    "other_support_needed": ("какой поддержки", "еще хотел", "фонда"),
    "had_tutor": ("репетитор",),
}

INTERVIEW_SUMMARY_MARKERS = {
    "interviewer_summary": ("комментарий",),
    "interviewer_risks": ("сведения для дополнительного контроля",),
    "interviewer_recommendations": ("нужна другая поддержка",),
    "interviewer_score": ("итоговая оценка",),
}

C_VALUE_MARKERS = (
    (("школа кандидата", "номер"), "Наименование и номер школы"),
    (("тип школы",), "Наименование и номер школы"),
    (("как далеко школа",), "Город_область_регион"),
    (("есть ли у школы специализация",), "Профиль_класса_при_наличии"),
    (("какой профиль класса",), "Профиль_класса_при_наличии"),
    (("по каким предметам", "егэ"), "Предметы_по_которым_планируешь_сдавать_ЕГЭ"),
    (("поступать по егэ", "олимпиадам"), "Ты_планируешь_поступать_по_ЕГЭ_или_олимпиадам_"),
    (("приоритет", "вуз"), "Список_наиболее_приоритетных_ВУЗов_в_которые_планируешь_поступать_не_более_5"),
    (("специальност", "поступлен"), "Список_специальностей_которые_рассматриваешь_для_поступления"),
    (("участв", "олимпиад"), "Планируешь_ли_ты_участвовать_в_олимпиадах"),
    (("дополнительная поддержка", "фонд"), "Выбранные курсы из видеовизитки"),
    (("какой поддержки", "фонд"), "Какую_помощь_ты_ожидаешь_от_Фонда"),
    (("достижен", "рассказать"), "Кратко_опиши_свои_достижения_за_последние_два_года"),
    (("достижен", "последние"), "Кратко_опиши_свои_достижения_за_последние_два_года"),
    (("работа родителей", "доход"), "Среднемесячный_доход_на_1_члена_семьи_за_последние_12_месяцев_руб_"),
    (("подготовк", "после школы"), "Как_ты_планируешь_свою_подготовку_к_поступлению_на_следующий_год_"),
    (("материальное положение",), "Другое важное"),
    (("компьютер", "интернет"), "Есть_ли_у_тебя_дома_компьютер_с_доступом_в_интернет"),
    (("выписки", "мотписьм", "видеовизит"), "Выписки из мотивационного письма"),
    (("мотивацион", "письм"), "Выписки из мотивационного письма"),
    (("видеовизит",), "Отзыв по видеовизитке"),
    (("выбранные", "курсы"), "Выбранные курсы из видеовизитки"),
)


def _application_value_coord(label: str, row: int) -> str:
    if _label_key(label) in {_label_key(item) for item in BELOW_VALUE_LABELS}:
        return f"H{row + 1}"
    return f"I{row}"


def _value_for_c_label(label: str, values: dict[str, str], normalized_values: dict[str, str]) -> str:
    raw_label = _text(label)
    if raw_label.lstrip().startswith(('"', "'")):
        return ""
    label_key = _label_key(label)
    if not label_key:
        return ""
    if label_key in normalized_values:
        return normalized_values[label_key]

    for markers, value_label in C_VALUE_MARKERS:
        marker_keys = tuple(_label_key(marker) for marker in markers)
        if all(marker in label_key for marker in marker_keys):
            return normalized_values.get(_label_key(value_label), "")

    label_words = _label_words(label)
    if len(label_words) < 2:
        return ""

    best_score = 0
    best_value = ""
    for value_label, value in values.items():
        value_text = _text(value)
        if not value_text:
            continue
        value_words = _label_words(value_label)
        if not value_words:
            continue
        overlap = label_words & value_words
        if not overlap:
            continue
        direct_match = _label_key(value_label) in label_key or label_key in _label_key(value_label)
        if len(overlap) < 2 and not direct_match:
            continue
        if "работа" in label_words and "работа" not in value_words and not direct_match:
            continue
        score = len(overlap) * 3
        if direct_match:
            score += 5
        score -= max(len(value_words - label_words), 0)
        if score > best_score:
            best_score = score
            best_value = value_text

    return best_value if best_score >= 5 else ""
    return ""


def _find_row_by_markers(ws, markers: tuple[str, ...], columns=("B", "C", "D", "H")) -> int | None:
    marker_keys = tuple(_label_key(marker) for marker in markers)
    for row in range(1, ws.max_row + 1):
        haystack = _label_key(" ".join(_text(ws[f"{col}{row}"].value) for col in columns))
        if haystack and all(marker in haystack for marker in marker_keys):
            return row
    return None


def _first_value(ws, coords: tuple[str, ...]):
    for coord in coords:
        value = _safe_get(ws, coord)
        if _text(value):
            return value
    return None


def _write_interview_result_values(ws, result):
    used_rows: set[int] = set()
    found_fields: set[str] = set()
    for field_name, markers in INTERVIEW_RESULT_MARKERS.items():
        row = _find_row_by_markers(ws, markers, columns=("B", "C"))
        if row and row not in used_rows:
            _safe_set(ws, f"D{row}", _field(result, field_name))
            _safe_set(ws, f"E{row}", "")
            used_rows.add(row)
            found_fields.add(field_name)

    for row, field_name in INTERVIEW_RESULT_ROWS.items():
        if field_name not in found_fields and row <= ws.max_row and row not in used_rows and _field(result, field_name):
            _safe_set(ws, f"D{row}", _field(result, field_name))
            _safe_set(ws, f"E{row}", "")
            used_rows.add(row)

    summary_rows = {
        field_name: _find_row_by_markers(ws, markers, columns=("C",))
        for field_name, markers in INTERVIEW_SUMMARY_MARKERS.items()
    }
    if any(summary_rows.values()):
        for field_name, row in summary_rows.items():
            if row:
                _safe_set(ws, f"F{row}", _field(result, field_name))
    else:
        _safe_set(ws, "F255", _field(result, "interviewer_summary"))
        _safe_set(ws, "F256", _field(result, "interviewer_risks"))
        _safe_set(ws, "F257", _field(result, "interviewer_recommendations"))
        _safe_set(ws, "F258", _field(result, "interviewer_score"))


def _normalize_for_model_field(field: models.Field, raw):
    if raw is None:
        return None
    if isinstance(raw, str) and raw.strip() == "":
        return None
    if isinstance(field, models.BooleanField):
        if isinstance(raw, bool):
            return raw
        text = str(raw).strip().lower()
        if text in {"да", "yes", "y", "true", "1", "истина"}:
            return True
        if text in {"нет", "no", "n", "false", "0", "ложь"}:
            return False
        return None
    if isinstance(field, (models.IntegerField, models.PositiveIntegerField, models.BigIntegerField, models.SmallIntegerField)):
        if isinstance(raw, int):
            return raw
        text = str(raw).strip()
        digits = "".join(ch for ch in text if ch.isdigit())
        if not digits:
            return None
        return (-1 if text.startswith("-") else 1) * int(digits)
    if isinstance(field, models.DecimalField):
        if isinstance(raw, Decimal):
            return raw
        cleaned = "".join(ch for ch in str(raw).strip().replace(",", ".") if ch in "0123456789.-")
        if cleaned in {"", "-", ".", "-."}:
            return None
        try:
            return Decimal(cleaned)
        except InvalidOperation:
            return None
    if isinstance(field, models.DateField) and not isinstance(field, models.DateTimeField):
        if hasattr(raw, "date"):
            return raw.date()
        return parse_date(str(raw).strip())
    if isinstance(field, models.DateTimeField):
        if hasattr(raw, "tzinfo"):
            return raw
        return parse_datetime(str(raw).strip())
    return str(raw).strip()


def _interview_result_fields(result_obj) -> list[models.Field]:
    skip = {"id", "pk", "interview", "created_at", "updated_at", "status"}
    return [
        field
        for field in result_obj._meta.get_fields()
        if getattr(field, "concrete", False)
        and not getattr(field, "auto_created", False)
        and field.name not in skip
    ]


def _parse_known_interview_result_values(ws) -> dict[str, object]:
    values: dict[str, object] = {}
    used_rows: set[int] = set()

    for field_name, markers in INTERVIEW_RESULT_MARKERS.items():
        row = _find_row_by_markers(ws, markers, columns=("B", "C"))
        if row and row not in used_rows:
            value = _first_value(ws, (f"E{row}", f"D{row}"))
            if _text(value):
                values[field_name] = value
                used_rows.add(row)

    for row, field_name in INTERVIEW_RESULT_ROWS.items():
        if field_name in values or row > ws.max_row or row in used_rows:
            continue
        value = _first_value(ws, (f"E{row}", f"D{row}"))
        if _text(value):
            values[field_name] = value
            used_rows.add(row)

    summary_rows = {
        field_name: _find_row_by_markers(ws, markers, columns=("C",))
        for field_name, markers in INTERVIEW_SUMMARY_MARKERS.items()
    }
    if any(summary_rows.values()):
        for field_name, row in summary_rows.items():
            if row:
                value = _first_value(ws, (f"F{row}", f"G{row}", f"D{row}", f"E{row}"))
                if _text(value):
                    values[field_name] = value
    else:
        for coord, field_name in (
            ("F255", "interviewer_summary"),
            ("F256", "interviewer_risks"),
            ("F257", "interviewer_recommendations"),
            ("F258", "interviewer_score"),
        ):
            value = _safe_get(ws, coord)
            if _text(value):
                values[field_name] = value

    return values


def _parse_verbose_interview_result_values(ws, result_obj) -> dict[str, object]:
    fields_by_key = {
        _label_key(getattr(field, "verbose_name", "") or field.name): field.name
        for field in _interview_result_fields(result_obj)
    }
    values: dict[str, object] = {}

    for row in range(1, ws.max_row + 1):
        label_candidates = (
            _text(_safe_get(ws, f"B{row}")),
            _text(_safe_get(ws, f"C{row}")),
            _text(_safe_get(ws, f"H{row}")),
        )
        for label in label_candidates:
            field_name = fields_by_key.get(_label_key(label))
            if not field_name or field_name in values:
                continue
            value = _first_value(ws, (f"E{row}", f"D{row}", f"F{row}", f"I{row}", f"H{row + 1}"))
            if _text(value):
                values[field_name] = value
    return values


def import_interview_result_xlsx(file_obj, result_obj) -> list[str]:
    workbook = load_workbook(file_obj, data_only=True)
    ws = workbook.active

    raw_values = {
        **_parse_verbose_interview_result_values(ws, result_obj),
        **_parse_known_interview_result_values(ws),
    }
    fields = {field.name: field for field in _interview_result_fields(result_obj)}
    updated_fields: list[str] = []

    for field_name, raw_value in raw_values.items():
        field = fields.get(field_name)
        if not field:
            continue
        value = _normalize_for_model_field(field, raw_value)
        if value is None:
            continue
        setattr(result_obj, field_name, value)
        updated_fields.append(field_name)

    if updated_fields:
        result_obj.save(update_fields=updated_fields + ["updated_at"])
    return updated_fields


def build_prefilled_interview_xlsx(user, interview) -> bytes:
    workbook = load_workbook(_template_path())
    ws = workbook.active

    _safe_set(ws, "D6", user.get_full_name() or user.username)
    _safe_set(ws, "D8", "")
    _safe_set(ws, "D9", timezone.localdate().strftime("%d.%m.%Y"))

    values = _build_application_values(user, interview)
    normalized_values = {_label_key(label): value for label, value in values.items()}
    for row in range(1, ws.max_row + 1):
        label = _text(ws[f"H{row}"].value)
        if _label_key(label) in normalized_values:
            _safe_set(ws, _application_value_coord(label, row), normalized_values[_label_key(label)])

        c_value = _value_for_c_label(ws[f"C{row}"].value, values, normalized_values)
        if c_value and not _text(_safe_get(ws, f"D{row}")):
            _safe_set(ws, f"D{row}", c_value)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
